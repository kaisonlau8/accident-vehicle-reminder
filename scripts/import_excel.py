"""DMS Excel 导入 → 标准化 repair_orders JSON。

DMS 导出的"维修工单总部查询"Excel 包含 4 个 Sheet：
- 维修工单（主表）：24 列，含门店/区域/状态/时间/VIN 等
- 工单工时：维修内容与类型
- 工单备件：配件明细
- 工单其他项目：其他费用

本项目的 DMS 爬虫在导出前已将业务类型固定为“事故维修”，因此这里直接
提取"维修工单"Sheet 并标准化输出，不再做事故车二次验证或过滤。
"""

import argparse
import json
import sys
from datetime import datetime, date, timedelta
from pathlib import Path

import openpyxl
from time_utils import beijing_today, beijing_iso

DATA_DIR = Path(__file__).resolve().parent.parent / "data" / "repair_orders"

IN_PROGRESS_STATUSES = {"待派工", "接车", "维修中"}
CANCELLED_STATUSES = {"已作废"}


def _parse_date_only(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    return str(val).strip()[:10] or None


def _days_between(date_str1: str, date_str2: str) -> int:
    try:
        d1 = datetime.strptime(date_str1[:10], "%Y-%m-%d").date()
        d2 = datetime.strptime(date_str2[:10], "%Y-%m-%d").date()
        return (d2 - d1).days
    except (ValueError, TypeError):
        return -1


def _is_cancelled_status(status: str) -> bool:
    return status in CANCELLED_STATUSES


def _is_in_progress_status(status: str) -> bool:
    return status in IN_PROGRESS_STATUSES


def _normalize_current_stage(status: str) -> str:
    """标准化当前阶段。

    规则：
    - 待派工 / 接车 / 维修中 -> 视为维修中状态，保留原状态名
    - 已作废 -> 已作废（理论上会在导入前被过滤）
    - 其他所有状态 -> 维修完成
    """
    if _is_cancelled_status(status):
        return "已作废"
    if _is_in_progress_status(status):
        return status
    return "维修完成"


def _is_qc_completed(status: str) -> bool:
    """判断是否已完成维修。"""
    return not _is_in_progress_status(status)


def import_excel(xlsx_path: str) -> dict:
    """导入 DMS 维修工单 Excel，返回标准 JSON。"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # 读取维修工单主表
    main_sheet = wb["维修工单"] if "维修工单" in wb.sheetnames else wb[wb.sheetnames[0]]

    # 读取表头
    headers = [str(cell.value or "") for cell in main_sheet[1]]
    col_idx = {h: i for i, h in enumerate(headers)}

    today = beijing_today().strftime("%Y-%m-%d")
    records = []

    for row in main_sheet.iter_rows(min_row=2, values_only=True):
        row_dict = {h: (row[i] if i < len(row) else None) for h, i in col_idx.items()}
        work_order_no = str(row_dict.get("派工单号", "") or "").strip()
        if not work_order_no:
            continue

        entry_date = _parse_date_only(row_dict.get("到店时间"))
        days_in_shop = _days_between(entry_date, today) if entry_date else -1
        status = str(row_dict.get("维修状态", "") or "").strip()
        if _is_cancelled_status(status):
            continue

        current_stage = _normalize_current_stage(status)

        record = {
            "repair_order_no": work_order_no,
            "store_code": str(row_dict.get("门店编码", "") or "").strip(),
            "store_name": str(row_dict.get("门店名称", "") or "").strip(),
            "region": str(row_dict.get("大区", "") or "").strip(),
            "status": status,
            "current_stage": current_stage,
            "service_advisor": str(row_dict.get("服务管家", "") or "").strip(),
            "entry_date": entry_date,
            "est_delivery_date": _parse_date_only(row_dict.get("预计交车时间")),
            "qc_date": _parse_date_only(row_dict.get("质检时间")),
            "settlement_date": _parse_date_only(row_dict.get("结算时间")),
            "payment_date": _parse_date_only(row_dict.get("收款时间")),
            "departure_date": _parse_date_only(row_dict.get("离店时间")),
            "plate_no": str(row_dict.get("车牌号", "") or "").strip(),
            "vin": str(row_dict.get("VIN码", "") or "").strip(),
            "vehicle_model": str(row_dict.get("车系名称", "") or "").strip(),
            "days_in_shop": days_in_shop,
            "is_qc_completed": _is_qc_completed(status),
            "sender_name": str(row_dict.get("送修人", "") or "").strip(),
            "sender_phone": str(row_dict.get("送修人电话", "") or "").strip(),
        }
        records.append(record)

    # 从实际数据推断日期范围（而非硬编码30天）
    entry_dates = [r.get("entry_date") for r in records if r.get("entry_date")]
    if entry_dates:
        actual_start = min(entry_dates)
        actual_end = max(entry_dates)
    else:
        actual_start = (beijing_today() - timedelta(days=30)).strftime("%Y-%m-%d")
        actual_end = today

    result = {
        "snapshot_date": today,
        "snapshot_start": actual_start,
        "snapshot_end": actual_end,
        "source": "dms_excel",
        "source_file": Path(xlsx_path).name,
        "generated_at": beijing_iso(),
        "total_records": len(records),
        "accident_records": len(records),
        "skipped_non_accident": 0,
        "records": records,
    }

    # 保存 JSON
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    output_path = DATA_DIR / f"repair_orders_{today}.json"

    # Sanity check: warn if data coverage looks suspiciously narrow
    if actual_start == actual_end and len(records) < 50:
        print(f"[WARN] 数据仅覆盖单日 ({actual_start})，共 {len(records)} 条记录")
        print("  可能爬取时日期范围未被正确设置，建议检查爬取日志")

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2, default=str)

    print(f"导入完成: {len(records)} 条记录")
    print(f"  事故车: {len(records)} 条 | 跳过非事故: 0 条")
    print(f"  输出: {output_path}")
    return result
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="导入 DMS 维修工单 Excel")
    parser.add_argument("xlsx_path", help="DMS 导出的维修工单 Excel 文件路径")
    args = parser.parse_args()

    if not Path(args.xlsx_path).exists():
        print(f"文件不存在: {args.xlsx_path}")
        sys.exit(1)

    import_excel(args.xlsx_path)
