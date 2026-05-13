"""报表生成器 — 生成门店/区域/全国 Excel 报表。"""

import json
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

REPORTS_DIR = Path(__file__).resolve().parent.parent / "data" / "reports"


def _style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")


def generate_national_report(rule_results: dict) -> str:
    """生成全国每日报表 Excel，返回文件路径。"""
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    snapshot = rule_results["snapshot_date"]
    national = rule_results["national_report"]

    wb = Workbook()

    # ── Sheet 1: 全国汇总 ──────────────────────────────────
    ws = wb.active
    ws.title = "全国汇总"
    ws.append(["指标", "数值"])
    ws.append(["在修车辆总数", national["total_vehicles"]])
    ws.append(["7~10天未完工", national["count_7d"]])
    ws.append(["10~14天未完工", national["count_10d"]])
    ws.append(["≥14天未完工", national["count_14d"]])
    ws.append(["30天内事故工单7天完工率", national["kpi_7d_rate"]])
    ws.append(["30天内事故工单10天完工率", national["kpi_10d_rate"]])
    ws.append(["数据时效", national.get("date_range", snapshot)])
    ws.append(["数据截止", snapshot])
    _style_header(ws)

    # ── Sheet 2: 区域汇总 ──────────────────────────────────
    ws2 = wb.create_sheet("区域汇总")
    ws2.append(["区域", "7~10天未完工", "10~14天未完工", "≥14天未完工"])
    for r in national.get("regions", []):
        ws2.append([r["region"], r["count_7d"], r["count_10d"], r["count_14d"]])
    _style_header(ws2)

    # ── Sheet 3: 超期车辆明细 ──────────────────────────────
    ws3 = wb.create_sheet("超期车辆明细")
    ws3.append(["门店编码", "门店名称", "区域", "VIN", "车型", "进店时间", "在店天数", "当前节点", "告警级别"])
    for alert in rule_results.get("alerts", []):
        level = alert.get("level", {})
        level_label = level.get("label", "") if level else ""
        ws3.append([
            alert["store_code"], alert["store_name"], alert["region"],
            alert["vin"], alert.get("vehicle_model", ""), alert.get("entry_date", ""),
            alert["days_in_shop"], alert["current_stage"], level_label,
        ])
    _style_header(ws3)

    # ── Sheet 4: 各门店汇总 ────────────────────────────────
    ws4 = wb.create_sheet("各门店汇总")
    ws4.append(["门店编码", "门店名称", "区域", "在修总数", "7~10天未完工", "10~14天未完工", "≥14天未完工"])
    for sc, sr in rule_results.get("store_reports", {}).items():
        ws4.append([sc, sr["store_name"], sr["region"], sr["total"], sr["count_7d"], sr["count_10d"], sr["count_14d"]])
    _style_header(ws4)

    output_path = REPORTS_DIR / f"national_report_{snapshot}.xlsx"
    wb.save(output_path)
    print(f"全国报表已生成: {output_path}")
    return str(output_path)


def generate_region_report_excel(region_name: str, region_data: dict, rule_results: dict) -> str:
    """生成区域报表 Excel，返回文件路径。"""
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    snapshot_date = region_data.get("snapshot_date", date.today().strftime("%Y-%m-%d"))

    wb = Workbook()

    # ── Sheet 1: 区域汇总 ──────────────────────────────────
    ws = wb.active
    ws.title = f"{region_name}汇总"
    ws.append(["门店名称", "在修车辆", "7~10天未完工", "10~14天未完工", "≥14天未完工"])
    for s in region_data.get("stores_summary", []):
        ws.append([s["store_name"], s.get("total", ""), s["count_7d"], s["count_10d"], s["count_14d"]])
    _style_header(ws)

    # ── Sheet 2: 超期车辆明细 ──────────────────────────────
    ws2 = wb.create_sheet("超期车辆明细")
    ws2.append(["门店名称", "VIN", "车型", "进店时间", "在店天数", "当前节点", "告警级别"])
    for alert in rule_results.get("alerts", []):
        if alert.get("region") != region_name:
            continue
        level = alert.get("level", {})
        level_label = level.get("label", "") if level else ""
        ws2.append([
            alert["store_name"], alert["vin"], alert.get("vehicle_model", ""),
            alert.get("entry_date", ""), alert["days_in_shop"],
            alert["current_stage"], level_label,
        ])
    _style_header(ws2)

    output_path = REPORTS_DIR / f"region_{region_name}_{snapshot_date}.xlsx"
    wb.save(output_path)
    print(f"区域报表已生成: {output_path}")
    return str(output_path)


def generate_all_region_reports(rule_results: dict) -> dict[str, str]:
    """为每个区域生成 Excel 报表，返回 {region_name: excel_path} 字典。"""
    excel_paths = {}
    for region_name, region_data in rule_results.get("region_reports", {}).items():
        if not region_data.get("count_7d", 0) and not region_data.get("count_10d", 0) and not region_data.get("count_14d", 0):
            continue
        excel_paths[region_name] = generate_region_report_excel(region_name, region_data, rule_results)
    return excel_paths


def generate_store_report_excel(store_code: str, store_data: dict, rule_results: dict) -> str:
    """生成门店报表 Excel，返回文件路径。"""
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    snapshot_date = store_data.get("snapshot_date", date.today().strftime("%Y-%m-%d"))
    store_name = store_data.get("store_name", store_code)

    wb = Workbook()

    # ── Sheet 1: 门店汇总 ──────────────────────────────────
    ws = wb.active
    ws.title = f"{store_name}汇总"
    ws.append(["指标", "数值"])
    ws.append(["在修车辆总数", store_data.get("total", 0)])
    ws.append(["7~10天未完工", store_data.get("count_7d", 0)])
    ws.append(["10~14天未完工", store_data.get("count_10d", 0)])
    ws.append(["≥14天未完工", store_data.get("count_14d", 0)])
    ws.append(["数据时效", store_data.get("date_range", snapshot_date)])
    ws.append(["数据截止", snapshot_date])
    _style_header(ws)

    # ── Sheet 2: 超期车辆明细 ──────────────────────────────
    ws2 = wb.create_sheet("超期车辆明细")
    ws2.append(["VIN", "车型", "进店时间", "在店天数", "当前节点", "告警级别"])
    for v in store_data.get("overdue_vehicles", []):
        level = v.get("level")
        level_label = level.get("label", "") if level else ""
        ws2.append([
            v.get("vin", ""), v.get("vehicle_model", ""),
            v.get("entry_date", ""), v.get("days_in_shop", 0),
            v.get("current_stage", ""), level_label,
        ])
    _style_header(ws2)

    output_path = REPORTS_DIR / f"store_{store_code}_{snapshot_date}.xlsx"
    wb.save(output_path)
    return str(output_path)


def generate_all_store_reports(rule_results: dict) -> dict[str, str]:
    """为每个门店生成 Excel 报表，返回 {store_code: excel_path} 字典。"""
    excel_paths = {}
    for store_code, store_data in rule_results.get("store_reports", {}).items():
        if not store_data.get("count_7d", 0) and not store_data.get("count_10d", 0) and not store_data.get("count_14d", 0):
            continue
        path = generate_store_report_excel(store_code, store_data, rule_results)
        print(f"门店报表已生成: {path}")
        excel_paths[store_code] = path
    return excel_paths


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("用法: python report_generator.py <rule_results.json>")
        sys.exit(1)

    with open(sys.argv[1], "r", encoding="utf-8") as f:
        data = json.load(f)

    path = generate_national_report(data)
    print(f"报表文件: {path}")
