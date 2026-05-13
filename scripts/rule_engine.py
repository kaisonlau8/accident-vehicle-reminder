"""规则引擎 — 评估提醒规则，产出通知列表。

规则 1: 事故车超期分级告警（≥7天黄色/≥10天橙色/≥14天红色，每日持续跟踪）
规则 2: 已作废工单不计入告警和统计
规则 3: 门店每日报表
规则 4: 区域每日报表
规则 5: 全国每日报表+KPI
"""

import json
from collections import defaultdict
from datetime import date, timedelta
from pathlib import Path

import openpyxl

CONFIG_DIR = Path(__file__).resolve().parent.parent / "config"
FOLLOWUP_DIR = Path(__file__).resolve().parent.parent / "data" / "followup"
ALERT_HISTORY_PATH = FOLLOWUP_DIR / "alert_history.json"
LIST_XLSX_PATH = Path(__file__).resolve().parent.parent / "list.xlsx"


def _load_alert_history() -> dict:
    """加载告警历史（VIN → {first_alert_date, alert_count}）。"""
    if ALERT_HISTORY_PATH.exists():
        with open(ALERT_HISTORY_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def _save_alert_history(history: dict):
    """保存告警历史。"""
    FOLLOWUP_DIR.mkdir(parents=True, exist_ok=True)
    with open(ALERT_HISTORY_PATH, "w", encoding="utf-8") as f:
        json.dump(history, f, ensure_ascii=False, indent=2)


def load_stores_config() -> dict:
    """从 list.xlsx 读取门店→区域→人员映射，生成与原 stores.json 相同结构。

    Excel 格式：
    序号 | 门店编码 | 区域 | 门店名称 | 督导姓名 | 督导飞书电话号
         | 提醒人1 | 电话1 | 提醒人2 | 电话2 | 提醒人3 | 电话3 | 提醒人4 | 电话4
    """
    if not LIST_XLSX_PATH.exists():
        return {"stores": {}, "regions": {}, "national_recipients": {}}

    wb = openpyxl.load_workbook(str(LIST_XLSX_PATH), data_only=True)
    ws = wb[wb.sheetnames[0]]

    stores = {}
    regions = {}
    supervisors_seen = {}  # 督导姓名 → {name, phone}

    for row in ws.iter_rows(min_row=2, values_only=True):
        store_code = str(row[1] or "").strip()
        if not store_code:
            continue

        region_name = str(row[2] or "").strip()
        store_name = str(row[3] or "").strip()
        sup_name = str(row[4] or "").strip()
        sup_phone = str(row[5] or "").strip() if row[5] else ""

        # 提醒人1~4
        remind_phones = []
        for i in [7, 9, 11, 13]:
            phone = str(row[i] or "").strip() if row[i] else ""
            if phone:
                remind_phones.append(phone)

        stores[store_code] = {
            "name": store_name,
            "region": region_name,
            "service_manager": {"name": "", "phone": remind_phones[0] if remind_phones else ""},
            "supervisor": {"name": sup_name, "phone": sup_phone},
            "remind_phones": remind_phones,
        }

        # 区域信息
        if region_name and sup_name:
            supervisors_seen.setdefault(sup_name, {"name": sup_name, "phone": sup_phone})
            regions[region_name] = {"name": region_name, "supervisor": supervisors_seen[sup_name]}

    # national_recipients: 从 stores.json 读取，由控制台维护
    nr = {}
    stores_json_path = CONFIG_DIR / "stores.json"
    if stores_json_path.exists():
        with open(stores_json_path, "r", encoding="utf-8") as f:
            legacy = json.load(f)
            nr = legacy.get("national_recipients", {})
            if isinstance(nr, list):
                nr = {}

    return {"stores": stores, "regions": regions, "national_recipients": nr}


def load_rules_config() -> list[dict]:
    path = CONFIG_DIR / "rules.json"
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)["rules"]


def _get_recipient_phones(rule: dict, store_config: dict, store_code: str) -> list[str]:
    """获取单条告警收件人手机号。

    Rule 1 (告警): 只通知门店提醒人（提醒人1~4）
    Rule 3 (门店报表): 只通知门店提醒人1（门店服务经理）
    """
    phones = []
    store_info = store_config.get("stores", {}).get(store_code, {})

    for role in rule.get("recipients", []):
        if role == "store_service_manager":
            # 门店报表只通知提醒人1
            mgr = store_info.get("service_manager", {})
            if mgr.get("phone"):
                phones.append(mgr["phone"])
        elif role == "remind_all":
            # 告警通知所有提醒人
            for p in store_info.get("remind_phones", []):
                if p:
                    phones.append(p)
        elif role == "supervisor":
            region_name = store_info.get("region", "")
            region_info = store_config.get("regions", {}).get(region_name, {})
            sup = region_info.get("supervisor", {})
            if sup.get("phone"):
                phones.append(sup["phone"])
        else:
            nr = store_config.get("national_recipients", {})
            phone = nr.get(role, "") if isinstance(nr, dict) else ""
            if phone:
                phones.append(phone)
    return list(set(phones))


def _get_all_national_recipient_phones(store_config: dict) -> list[str]:
    """获取控制台配置的全部全国收件人手机号。"""
    nr = store_config.get("national_recipients", {})
    if not isinstance(nr, dict):
        return []
    phones = []
    for phone in nr.values():
        phone = str(phone or "").strip()
        if phone:
            phones.append(phone)
    return list(set(phones))


def _get_region_recipient_phones(region_name: str, store_config: dict) -> list[str]:
    """获取区域报表收件人手机号（该区域督导+庄帅+杨永昌）。"""
    phones = []
    region_info = store_config.get("regions", {}).get(region_name, {})
    sup = region_info.get("supervisor", {})
    if sup.get("phone"):
        phones.append(sup["phone"])
    phones.extend(_get_all_national_recipient_phones(store_config))
    return list(set(phones))


def _get_national_recipient_phones(store_config: dict) -> list[str]:
    """获取全国报表收件人手机号（全部督导+控制台配置的全国收件人）。"""
    phones = []
    for region_name, region_info in store_config.get("regions", {}).items():
        sup = region_info.get("supervisor", {})
        if sup.get("phone"):
            phones.append(sup["phone"])
    phones.extend(_get_all_national_recipient_phones(store_config))
    return list(set(phones))


def _determine_alert_level(days: int, levels: list[dict]) -> dict | None:
    """根据进店天数确定告警级别，取满足的最高阈值。"""
    for lv in reversed(levels):  # 从14→10→7检查
        if days >= lv["threshold"]:
            return lv
    return None


def _format_level_message(level: dict, rec: dict) -> str:
    """格式化分级告警文本消息。"""
    return (
        f"{level['icon']} {level['label']}\n"
        f"门店：{rec.get('store_name', '')}\n"
        f"VIN：{rec.get('vin', '')}\n"
        f"车型：{rec.get('vehicle_model', '')}\n"
        f"进店时间：{rec.get('entry_date', '')}\n"
        f"已进店 {rec.get('days_in_shop', 0)} 天未完工\n"
        f"当前节点：{rec.get('current_stage', '')}"
    )


def evaluate_rules(repair_orders: dict, stores_config: dict | None = None) -> dict:
    """评估所有规则，返回通知任务。"""
    if stores_config is None:
        stores_config = load_stores_config()

    rules = load_rules_config()
    records = repair_orders.get("records", [])
    snapshot_date = repair_orders.get("snapshot_date", date.today().strftime("%Y-%m-%d"))
    snapshot_start = repair_orders.get("snapshot_start", (date.today() - timedelta(days=30)).strftime("%Y-%m-%d"))
    date_range = f"{snapshot_start} ~ {snapshot_date}"

    results = {
        "snapshot_date": snapshot_date,
        "date_range": date_range,
        "alerts": [],
        "store_reports": {},
        "region_reports": {},
        "national_report": {},
    }

    # ── Rule 1: 分级告警（≥7天未完工，每日跟踪）───────────────
    alert_rule = None
    for rule in rules:
        if rule["id"] == 1:
            alert_rule = rule
            break

    alert_history = _load_alert_history()

    if alert_rule:
        levels = alert_rule.get("levels", [])

        for rec in records:
            days = rec.get("days_in_shop", -1)
            if days < 7 or rec.get("is_qc_completed") or rec.get("current_stage") == "已作废":
                continue

            level = _determine_alert_level(days, levels)
            if level is None:
                continue

            vin = rec.get("vin", "")
            store_code = rec.get("store_code", "")
            phones = _get_recipient_phones(alert_rule, stores_config, store_code)

            # 告警历史追踪
            hist_entry = alert_history.get(vin)
            if hist_entry:
                alert_count = hist_entry["alert_count"] + 1
                first_alert_date = hist_entry["first_alert_date"]
            else:
                alert_count = 1
                first_alert_date = snapshot_date
            alert_history[vin] = {"first_alert_date": first_alert_date, "alert_count": alert_count}

            msg = _format_level_message(level, rec)

            results["alerts"].append({
                "rule_id": 1,
                "rule_name": alert_rule["name"],
                "level": level,
                "store_code": store_code,
                "store_name": rec.get("store_name", ""),
                "region": rec.get("region", ""),
                "vin": vin,
                "vehicle_model": rec.get("vehicle_model", ""),
                "days_in_shop": days,
                "entry_date": rec.get("entry_date", ""),
                "current_stage": rec.get("current_stage", ""),
                "recipient_phones": phones,
                "msg_type": alert_rule["msg_type"],
                "message": msg,
                "alert_count": alert_count,
                "first_alert_date": first_alert_date,
            })

    _save_alert_history(alert_history)

    # ── Rule 3, 4, 5: 汇总报表 ────────────────────────────
    active_records = [r for r in records if not r.get("is_qc_completed") and r.get("current_stage") != "已作废"]

    # 按门店分组
    store_groups = defaultdict(list)
    for rec in active_records:
        store_groups[rec.get("store_code", "UNKNOWN")].append(rec)

    # 按区域分组
    region_groups = defaultdict(list)
    for rec in active_records:
        region_name = rec.get("region", "") or "未分配区域"
        region_groups[region_name].append(rec)

    # Rule 3: 门店报表
    for store_code, vehicles in store_groups.items():
        store_info = stores_config.get("stores", {}).get(store_code, {})
        count_7d = sum(1 for v in vehicles if 7 <= v.get("days_in_shop", 0) < 10)
        count_10d = sum(1 for v in vehicles if 10 <= v.get("days_in_shop", 0) < 14)
        count_14d = sum(1 for v in vehicles if v.get("days_in_shop", 0) >= 14)

        overdue_vehicles = [
            {
                "vin": v.get("vin", ""),
                "days_in_shop": v.get("days_in_shop", 0),
                "current_stage": v.get("current_stage", ""),
                "entry_date": v.get("entry_date", ""),
                "vehicle_model": v.get("vehicle_model", ""),
                "level": _determine_alert_level(v.get("days_in_shop", 0), levels) if levels else None,
            }
            for v in vehicles if v.get("days_in_shop", 0) >= 7
        ]

        results["store_reports"][store_code] = {
            "store_name": store_info.get("name", vehicles[0].get("store_name", "")),
            "region": store_info.get("region", vehicles[0].get("region", "")),
            "total": len(vehicles),
            "count_7d": count_7d,
            "count_10d": count_10d,
            "count_14d": count_14d,
            "overdue_vehicles": overdue_vehicles,
            "snapshot_date": snapshot_date,
            "date_range": date_range,
            "recipient_phones": [
                store_info.get("service_manager", {}).get("phone", "")
            ] if store_info.get("service_manager", {}).get("phone") else [],
        }

    # Rule 4: 区域报表
    for region_name, vehicles in region_groups.items():
        region_info = stores_config.get("regions", {}).get(region_name, {})
        stores_summary = []
        seen_stores = set()
        for v in vehicles:
            sc = v.get("store_code", "")
            if sc in seen_stores:
                continue
            seen_stores.add(sc)
            sr = results["store_reports"].get(sc, {})
            stores_summary.append({
                "store_name": sr.get("store_name", v.get("store_name", "")),
                "total": sr.get("total", 0),
                "count_7d": sr.get("count_7d", 0),
                "count_10d": sr.get("count_10d", 0),
                "count_14d": sr.get("count_14d", 0),
            })

        results["region_reports"][region_name] = {
            "store_count": len(seen_stores),
            "total_vehicles": len(vehicles),
            "count_7d": sum(1 for v in vehicles if 7 <= v.get("days_in_shop", 0) < 10),
            "count_10d": sum(1 for v in vehicles if 10 <= v.get("days_in_shop", 0) < 14),
            "count_14d": sum(1 for v in vehicles if v.get("days_in_shop", 0) >= 14),
            "stores_summary": stores_summary,
            "snapshot_date": snapshot_date,
            "date_range": date_range,
            "recipient_phones": _get_region_recipient_phones(region_name, stores_config),
        }

    # Rule 5: 全国报表
    count_7d = sum(1 for v in active_records if 7 <= v.get("days_in_shop", 0) < 10)
    count_10d = sum(1 for v in active_records if 10 <= v.get("days_in_shop", 0) < 14)
    count_14d = sum(1 for v in active_records if v.get("days_in_shop", 0) >= 14)
    total = len(active_records)

    total_accident_orders_30d = len(records)
    completed = [r for r in records if r.get("is_qc_completed")]
    completed_7d = sum(1 for r in completed if r.get("days_in_shop", 0) <= 7)
    completed_10d = sum(1 for r in completed if r.get("days_in_shop", 0) <= 10)

    # KPI口径：
    # 30天内事故工单7天完工率 = 7天内完工事故车辆数 / 全部事故工单数 * 100%
    # 30天内事故工单10天完工率 = 10天内完工事故车辆数 / 全部事故工单数 * 100%
    denominator = max(total_accident_orders_30d, 1)
    kpi_7d_rate = f"{completed_7d / denominator * 100:.1f}%"
    kpi_10d_rate = f"{completed_10d / denominator * 100:.1f}%"

    regions_summary = []
    for rn, rv in results["region_reports"].items():
        regions_summary.append({
            "region": rn,
            "count_7d": rv["count_7d"],
            "count_10d": rv["count_10d"],
            "count_14d": rv["count_14d"],
        })

    results["national_report"] = {
        "total_vehicles": total,
        "count_7d": count_7d,
        "count_10d": count_10d,
        "count_14d": count_14d,
        "kpi_7d_rate": kpi_7d_rate,
        "kpi_10d_rate": kpi_10d_rate,
        "total_accident_orders_30d": total_accident_orders_30d,
        "completed_7d": completed_7d,
        "completed_10d": completed_10d,
        "regions": regions_summary,
        "snapshot_date": snapshot_date,
        "date_range": date_range,
        "recipient_phones": _get_national_recipient_phones(stores_config),
    }

    return results


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("用法: python rule_engine.py <repair_orders.json>")
        sys.exit(1)

    with open(sys.argv[1], "r", encoding="utf-8") as f:
        data = json.load(f)

    result = evaluate_rules(data)
    print(json.dumps({
        "alert_count": len(result["alerts"]),
        "store_report_count": len(result["store_reports"]),
        "region_report_count": len(result["region_reports"]),
        "national_summary": {
            "total": result["national_report"]["total_vehicles"],
            "7d": result["national_report"]["count_7d"],
            "10d": result["national_report"]["count_10d"],
            "14d": result["national_report"]["count_14d"],
            "kpi_7d": result["national_report"]["kpi_7d_rate"],
            "kpi_10d": result["national_report"]["kpi_10d_rate"],
            "date_range": result["national_report"]["date_range"],
        }
    }, ensure_ascii=False, indent=2))
